#!/usr/bin/env -S uv run --python-preference=managed
# /// script
# requires-python = ">=3.8"
# dependencies = [
#   "openpyxl>=3.1",
#   "textual>=0.60.0" 
# ]
# ///

"""
MakerSpace Sign-In app
TUI Edition (powered by Textual)
"""

# -----------------------
# Debug + error reporting
# -----------------------
import os, sys, traceback, re
from datetime import datetime

APP_TITLE = "MakerSpace Sign-In"
_DIR = os.path.dirname(os.path.abspath(__file__))
_ERR_LOG = os.path.join(_DIR, "tui_sign_error.log")

# Officer codes DB (must be in same folder)
OFFICER_CODES_FILE = os.path.join(_DIR, "officer_codes.txt")

# --- Textual Imports ---
from textual.app import App, ComposeResult
from textual.screen import Screen, ModalScreen
from textual.containers import Container, Horizontal, Vertical
from textual.widgets import (
    Header,
    Footer,
    Button,
    Static,
    Input,
    Label,
    DataTable,
)
from textual.validation import Validator, ValidationResult, Regex


def _show_windows_message_box(text: str, title: str) -> None:
    # Kept from original script for crash handling
    try:
        import ctypes
        ctypes.windll.user_32.MessageBoxW(0, text, title, 0x00000010)
    except Exception:
        pass


def _write_err_log(text: str) -> None:
    try:
        with open(_ERR_LOG, "a", encoding="utf-8") as f:
            f.write("\n" + "=" * 80 + "\n")
            f.write(text)
            f.write("\n")
    except Exception:
        pass


def _excepthook(exc_type, exc_value, exc_tb):
    # Kept from original script
    tb = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
    try:
        print(tb, file=sys.stderr, flush=True)
    except Exception:
        pass
    _write_err_log(tb)
    if sys.platform.startswith("win"):
        _show_windows_message_box(tb, f"{APP_TITLE} - Unhandled Error")
    sys.__excepthook__(exc_type, exc_value, exc_tb)


# Install global error handler
sys.excepthook = _excepthook


# -----------------------
# Core Application Logic
# -----------------------

# ### EMAIL USERNAME + FIXED DOMAIN
STUDENT_DOMAIN = "@student.monroecc.edu"
MAX_USERNAME_LEN = 60 - len(STUDENT_DOMAIN)

# ---- OFFICER MENU PIN (env overrides; default = 3132) ----
OFFICER_MENU_PIN = os.environ.get("ELC_OFFICER_PIN", "3132")


def get_daily_makerspace_log_path() -> str:
    """Gets path to today's log, creating it with headers if it doesn't exist."""
    filename = f"MakerSpace_Log_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    filepath = os.path.join(_DIR, filename)
    if not os.path.exists(filepath):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter as _gcl

            wb = Workbook()
            ws = wb.active
            ws.title = "Log"
            headers = [
                "First Name",
                "Last Name",
                "Email",
                "Sign-In Time",
                "Sign-Out Time",
                "Duration (Minutes)",
            ]

            thin = Side(style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            header_font = Font(bold=True)
            header_fill = PatternFill("solid", fgColor="FFF2CC")

            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.border = border
                cell.fill = header_fill
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            ws.column_dimensions[_gcl(1)].width = 20  # First
            ws.column_dimensions[_gcl(2)].width = 20  # Last
            ws.column_dimensions[_gcl(3)].width = 30  # Email
            ws.column_dimensions[_gcl(4)].width = 15  # Sign-In
            ws.column_dimensions[_gcl(5)].width = 15  # Sign-Out
            ws.column_dimensions[_gcl(6)].width = 20  # Duration

            ws.freeze_panes = ws.cell(row=2, column=1)
            wb.save(filepath)
            wb.close()
        except Exception as e:
            _write_err_log(f"Failed to create new MakerSpace log file: {e}")
            # In TUI, we notify the app screen
            return None  # Indicate failure
    return filepath


# --- Validation Logic ---
class NameValidator(Validator):
    def validate(self, value: str) -> ValidationResult:
        if len(value) > 50:
            return self.failure("Name is too long (max 50).")
        if not value.strip():
            return self.failure("Field cannot be empty.")
        return self.success()


class UsernameValidator(Validator):
    def validate(self, value: str) -> ValidationResult:
        if not value:
            return self.failure("Field cannot be empty.")
        if len(value) > MAX_USERNAME_LEN:
            return self.failure(f"Username is too long (max {MAX_USERNAME_LEN}).")
        if not re.match(r"^[a-z0-9.\-]+$", value):
            return self.failure("Use only letters, numbers, dot, or dash.")
        return self.success()


# ---------- Officer codes persistence (TXT in same folder) ----------
def _ensure_codes_file_exists():
    if os.path.exists(OFFICER_CODES_FILE):
        return
    try:
        with open(OFFICER_CODES_FILE, "w", encoding="utf-8") as f:
            f.write(
                f"# Officer codes file - keep this file in the SAME folder as the app\n"
            )
            f.write(
                "# One entry per line: CODE,Role  (comma or pipe '|' accepted)\n"
            )
            f.write("# Example: 1234,President\n")
            f.write("# Lines starting with # are ignored.\n")
    except Exception as e:
        _write_err_log(f"Failed to create officer codes file: {e}")


def load_officer_codes() -> dict:
    codes = {}
    try:
        _ensure_codes_file_exists()
        with open(OFFICER_CODES_FILE, "r", encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if not s or s.startswith("#"):
                    continue
                parts = s.split("|", 1) if "|" in s else s.split(",", 1)
                if len(parts) != 2:
                    continue
                code, role = parts[0].strip(), parts[1].strip()
                if code.isdigit() and len(code) == 4 and role:
                    codes[code] = role
    except Exception as e:
        _write_err_log(f"Failed to load officer codes: {e}")
    return codes


def save_officer_codes(codes: dict) -> None:
    # This function can raise errors, which we'll catch
    with open(OFFICER_CODES_FILE, "w", encoding="utf-8") as f:
        f.write(f"# Officer codes file - keep this file in the SAME folder as the app\n")
        f.write("# CODE,Role\n")
        for code, role in sorted(codes.items()):
            f.write(f"{code},{role}\n")


# ---------------------------------
# TUI Modal Screens (Dialogs)
# ---------------------------------

class PINScreen(ModalScreen[bool]):
    """Modal screen to get officer PIN."""
    
    CSS = """
    PINScreen {
        align: center middle;
    }
    
    #pin-dialog {
        width: 40;
        height: 10;
        background: $boost;
        border: $heavy;
        padding: 1 2;
    }
    
    #pin-status {
        width: 100%;
        text-align: center;
        color: #FF5858; /* $error */
        height: 1;
    }
    
    #pin-input {
        width: 100%;
        margin-top: 1;
    }
    """

    def compose(self) -> ComposeResult:
        with Vertical(id="pin-dialog"):
            yield Label("Enter 4-digit Officer Passcode:")
            yield Input(
                password=True,
                max_length=4,
                id="pin-input",
                validators=[Regex(r"^\d{4}$", "Must be 4 digits.")],
            )
            yield Static(id="pin-status")
            yield Button("Cancel", id="cancel")

    def on_mount(self) -> None:
        self.query_one(Input).focus()

    def on_input_submitted(self, event: Input.Submitted) -> None:
        if event.validation_result and event.validation_result.is_valid:
            if event.value == OFFICER_MENU_PIN:
                self.dismiss(True)
            else:
                self.query_one("#pin-status").update("Incorrect passcode.")
                self.query_one(Input).value = ""
        else:
            self.query_one("#pin-status").update("Must be 4 digits.")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "cancel":
            self.dismiss(False)


class OfficerManagerScreen(ModalScreen):
    """Modal screen to manage officer codes."""
    
    CSS = """
    OfficerManagerScreen {
        align: center middle;
    }
    
    #manager-dialog {
        width: 80%;
        height: 80%;
        max-width: 100;
        max-height: 30;
        background: $boost;
        border: $heavy;
    }
    
    #codes-table {
        height: 1fr;
        margin: 1;
        border: $solid $background-lighten-2;
    }
    
    #manager-form {
        height: auto;
        padding: 0 1;
    }
    
    #manager-buttons {
        height: auto;
        padding: 1;
    }
    """

    def compose(self) -> ComposeResult:
        with Vertical(id="manager-dialog"):
            yield Label("Officer Codes Manager", id="manager-title")
            yield DataTable(id="codes-table")
            with Horizontal(id="manager-form"):
                yield Input(placeholder="Code (4 digits)", id="code", max_length=4)
                yield Input(placeholder="Role", id="role", expand=True)
            with Horizontal(id="manager-buttons"):
                yield Button("Add/Update", id="add")
                yield Button("Delete", id="delete")
                yield Button("Reload", id="reload")
                yield Button("Save", id="save")
                yield Button("Close", id="close", variant="error")

    def on_mount(self) -> None:
        table = self.query_one(DataTable)
        table.add_columns("Code", "Role")
        table.cursor_type = "row"
        self.refresh_table()

    def refresh_table(self) -> None:
        table = self.query_one(DataTable)
        table.clear()
        for code, role in sorted(self.app.officer_codes.items()):
            table.add_row(code, role, key=code)

    def on_data_table_row_selected(self, event: DataTable.RowSelected) -> None:
        code = event.row_key.value
        if code in self.app.officer_codes:
            role = self.app.officer_codes[code]
            self.query_one("#code").value = code
            self.query_one("#role").value = role

    def on_button_pressed(self, event: Button.Pressed) -> None:
        code_input = self.query_one("#code")
        role_input = self.query_one("#role")
        code = code_input.value.strip()
        role = role_input.value.strip()

        if event.button.id == "add":
            if not (code.isdigit() and len(code) == 4):
                self.app.notify("Code must be 4 digits.", title="Error", severity="error")
                return
            if not role:
                self.app.notify("Role cannot be empty.", title="Error", severity="error")
                return
            self.app.officer_codes[code] = role
            self.refresh_table()
            code_input.value = ""
            role_input.value = ""
            self.app.notify(f"Added/Updated code {code}", title="Success")

        elif event.button.id == "delete":
            table = self.query_one(DataTable)
            if table.cursor_row < 0:
                 self.app.notify("Select a row to delete.", title="Error", severity="error")
                 return
            code_to_del = table.get_row_at(table.cursor_row)[0]
            if code_to_del in self.app.officer_codes:
                del self.app.officer_codes[code_to_del]
                self.refresh_table()
                self.app.notify(f"Deleted code {code_to_del}", title="Success")

        elif event.button.id == "reload":
            self.app.officer_codes = load_officer_codes()
            self.refresh_table()
            self.app.notify("Reloaded codes from file.", title="Success")

        elif event.button.id == "save":
            try:
                save_officer_codes(self.app.officer_codes)
                self.app.notify("Officer codes saved to file.", title="Success")
            except Exception as e:
                _write_err_log(f"Failed to save officer codes: {e}")
                self.app.notify(f"Save failed: {e}", title="Error", severity="error")

        elif event.button.id == "close":
            self.dismiss()


# ---------------------------------
# TUI Main Screen
# ---------------------------------

class MakerSpaceScreen(Screen):
    """Main screen for makerspace sign-in/out."""

    def compose(self) -> ComposeResult:
        yield Header()
        with Container(id="main-content"):
            with Vertical(id="left-panel"):
                yield Label("MakerSpace / Room Sign-In")
                
                yield Label("First Name:")
                yield Input(id="first", validators=[NameValidator()])
                yield Label("Last Name:")
                yield Input(id="last", validators=[NameValidator()])
                yield Label(f"Email Username: ( ...{STUDENT_DOMAIN} )")
                yield Input(id="username", validators=[UsernameValidator()])
                
                with Horizontal(id="makerspace-buttons"):
                    yield Button("Sign IN", id="signin", variant="success")
                    yield Button("Sign OUT", id="signout", variant="error")
                
                yield Static(id="status", classes="status-line")

            with Vertical(id="right-panel"):
                yield Label("Instructions:", classes="instructions-title")
                yield Static(
                    "• Signing IN:\n"
                    "  1. Fill your First, Last, and Email Username.\n"
                    "  2. Click 'Sign IN'.\n\n"
                    "• Signing OUT:\n"
                    "  1. Fill the *same* info (especially Email).\n"
                    "  2. Click 'Sign OUT'.\n\n"
                    "Your time will be logged automatically.",
                    classes="instructions"
                )
        yield Footer()

    def on_mount(self) -> None:
        self.app.title = APP_TITLE
        self.app.sub_title = ""
        self.query_one("#first").focus()
        self.query_one("#status").update("Welcome! Enter your info to sign in.")

    def _get_validated_input(self) -> tuple[str, str, str] | None:
        """Helper to get and validate all inputs."""
        status_label = self.query_one("#status")
        first_input = self.query_one("#first")
        last_input = self.query_one("#last")
        user_input = self.query_one("#username")

        first = first_input.value.strip()
        last = last_input.value.strip()
        username = user_input.value.strip().lower()
        
        if not all([
            first_input.validation_result, first_input.validation_result.is_valid,
            last_input.validation_result, last_input.validation_result.is_valid,
            user_input.validation_result, user_input.validation_result.is_valid
        ]):
            self.app.notify("Please fix all fields.", title="Error", severity="error")
            status_label.update("Please fill all fields correctly.")
            status_label.set_classes("bad")
            return None

        email = f"{username}{STUDENT_DOMAIN}"
        return (first, last, email)
    
    def _clear_fields(self):
        self.query_one("#first").value = ""
        self.query_one("#last").value = ""
        self.query_one("#username").value = ""
        self.query_one("#first").focus()

    def on_input_submitted(self, event: Input.Submitted) -> None:
        # Default <Enter> on last field to "Sign IN"
        if event.input.id == "username":
            self.on_button_pressed(Button.Pressed(self.query_one("#signin")))

    def on_button_pressed(self, event: Button.Pressed) -> None:
        inputs = self._get_validated_input()
        if not inputs:
            return

        first, last, email = inputs
        status_label = self.query_one("#status")
        
        if event.button.id == "signin":
            result, message = self.app.makerspace_sign_in(first, last, email)
            if result == "good":
                status_label.update(message)
                status_label.set_classes("good")
                self._clear_fields()
            else:
                status_label.update(message)
                status_label.set_classes("warn" if result == "warn" else "bad")

        elif event.button.id == "signout":
            result, message = self.app.makerspace_sign_out(first, last, email)
            if result == "good":
                status_label.update(message)
                status_label.set_classes("good")
                self._clear_fields()
            else:
                status_label.update(message)
                status_label.set_classes("warn" if result == "warn" else "bad")


# ---------------------------------
# The Main Textual App
# ---------------------------------

class ELCSignInApp(App):
    """The main TUI application class."""

    # --- Embed CSS ---
    # This version is simplified for MakerSpace-only and
    # uses NO CSS variables for colors to prevent errors.
    CSS = """
    Screen {
        background: #000000;
        color: #C99700;
    }
    
    Header {
        background: #C99700;
        color: #000000;
        text-style: bold;
    }
    
    Footer {
        background: #C99700;
        color: #000000;
    }
    
    Button {
        background: #D9D9D9;
        color: #111111;
        border: none;
    }
    Button:hover {
        background: #C0C0C0; /* Darker gray */
    }
    Button.-primary {
        background: #C99700;
        color: #000000;
    }
    Button.-success {
        background: #7DD97D;
        color: #000000;
    }
    Button.-error {
        background: #FF5858;
        color: #000000;
    }
    
    Input {
        background: #D9D9D9;
        color: #111111;
        border: none;
    }
    Input:focus {
        border: $solid #C99700;
    }
    Input.-invalid {
        border: $solid #FF5858;
    }
    
    Label {
        margin: 1 0 0 0;
    }
    
    #main-content {
        layout: horizontal;
        height: 1fr;
    }
    
    #left-panel {
        width: 1fr;
        padding: 1 2;
        border-right: $solid #C99700;
    }
    
    #right-panel {
        width: 1fr;
        padding: 1 2;
    }
    
    .status-line {
        height: 1;
        margin-top: 1;
    }
    .good { color: #7DD97D; }
    .bad { color: #FF5858; }
    .warn { color: #FFD16A; }
    
    /* MakerSpace Screen */
    #makerspace-buttons {
        margin-top: 1;
        height: auto;
    }
    
    .instructions-title {
        margin-top: 2;
        text-style: bold;
    }
    .instructions {
        color: #A07900; /* Darker gold */
    }
    
    /* Manager Screen */
    #manager-title {
        width: 100%;
        text-align: center;
        text-style: bold;
    }
    """

    # --- Key Bindings ---
    BINDINGS = [
        ("ctrl+q", "quit", "Quit"),
        ("ctrl+l", "toggle_lock", "Lock/Unlock"),
        ("ctrl+m", "manage_codes", "Manage Codes"),
    ]

    def __init__(self):
        super().__init__()
        # App-wide state
        self.officer_codes = load_officer_codes()
        self.officer_unlocked = False

    def on_mount(self) -> None:
        self.push_screen(MakerSpaceScreen())

    # --- Business Logic (Methods) ---

    def makerspace_sign_in(self, first: str, last: str, email: str) -> tuple[str, str]:
        """Logs a makerspace sign-in. Returns (status, message)."""
        now = datetime.now()
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "bad", "ERROR: openpyxl not found. Cannot write to log."

        filepath = get_daily_makerspace_log_path()
        if not filepath:
            return "bad", "ERROR: Could not create or access log file."

        try:
            wb = load_workbook(filepath)
            ws = wb.active
            for row in range(ws.max_row, 1, -1):
                if ws.cell(row, 3).value == email and ws.cell(row, 5).value is None:
                    wb.close()
                    return "warn", f"Already signed in (at {ws.cell(row, 4).value}). Sign out first."
            
            new_row = [first, last, email, now.strftime("%H:%M:%S"), None, None]
            ws.append(new_row)
            wb.save(filepath)
            wb.close()
            return "good", f"Signed In: {first} {last}"
        except PermissionError:
            return "bad", "ERROR: Log file is open in Excel. Please close it."
        except Exception as e:
            _write_err_log(f"MakerSpace Sign-In Error: {e}")
            return "bad", f"An error occurred: {e}"


    def makerspace_sign_out(self, first: str, last: str, email: str) -> tuple[str, str]:
        """Logs a makerspace sign-out. Returns (status, message)."""
        now = datetime.now()
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "bad", "ERROR: openpyxl not found. Cannot write to log."

        filepath = get_daily_makerspace_log_path()
        if not filepath or not os.path.exists(filepath):
            return "bad", "ERROR: No log file found. Cannot sign out."

        try:
            wb = load_workbook(filepath)
            ws = wb.active
            found_row_idx = -1
            # Find the user's open session
            # Iterate backwards
            for row_idx in range(ws.max_row, 1, -1):
                if ws.cell(row_idx, 3).value == email and ws.cell(row_idx, 5).value is None:
                    found_row_idx = row_idx
                    break
            
            if found_row_idx == -1:
                wb.close()
                return "warn", "Could not find an open sign-in for this email."

            sign_in_str = ws.cell(found_row_idx, 4).value
            sign_in_time = datetime.strptime(sign_in_str, "%H:%M:%S")
            sign_in_dt = now.replace(hour=sign_in_time.hour, minute=sign_in_time.minute, second=sign_in_time.second, microsecond=0)
            
            if now < sign_in_dt:
                duration_minutes = "N/A (Past Midnight)"
            else:
                duration = now - sign_in_dt
                duration_minutes = round(duration.total_seconds() / 60, 1)

            ws.cell(found_row_idx, 5, value=now.strftime("%H:%M:%S"))
            ws.cell(found_row_idx, 6, value=duration_minutes)
            
            wb.save(filepath)
            wb.close()
            return "good", f"Signed Out: {first} {last}. Duration: {duration_minutes} min."
        except PermissionError:
            return "bad", "ERROR: Log file is open in Excel. Please close it."
        except Exception as e:
            _write_err_log(f"MakerSpace Sign-Out Error: {e}")
            return "bad", f"An error occurred: {e}"

    # --- Action Handlers (from Bindings) ---

    def action_quit(self) -> None:
        """Quits the application."""
        self.exit()

    def _check_unlocked(self, feature_name: str) -> bool:
        """Utility to check for officer lock."""
        if not self.officer_unlocked:
            self.notify(
                f"You must unlock settings to {feature_name}.",
                title="Locked",
                severity="error",
            )
            return False
        return True

    async def action_toggle_lock(self) -> None:
        """Toggles the officer-locked state."""
        if self.officer_unlocked:
            self.officer_unlocked = False
            self.notify("Settings have been locked.", title="Locked")
        else:
            def after_pin(unlocked: bool):
                if unlocked:
                    self.officer_unlocked = True
                    self.notify("Settings unlocked.", title="Success")
                else:
                    self.notify("PIN entry cancelled.", title="Locked")
            
            await self.push_screen(PINScreen(), after_pin)

    async def action_manage_codes(self) -> None:
        """Opens the officer code manager."""
        if not self._check_unlocked("manage codes"):
            return
        await self.push_screen(OfficerManagerScreen())


if __name__ == "__main__":
    app = ELCSignInApp()
    app.run()