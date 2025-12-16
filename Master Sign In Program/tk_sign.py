#!/usr/bin/env -S uv run --python-preference=managed
# /// script
# requires-python = ">=3.8"
# dependencies = [
#   "openpyxl>=3.1"
# ]
# ///

"""
Dual-mode ELC Sign-In app (Event Kiosk and MakerSpace Logger)
"""

# -----------------------
# Debug + error reporting
# -----------------------
import os, sys, traceback, json # Added json

APP_TITLE = "ELC Sign-In"
_DIR = os.path.dirname(os.path.abspath(__file__))
_ERR_LOG = os.path.join(_DIR, "tk_sign_error.log")

# Officer codes DB (must be in same folder)
OFFICER_CODES_FILE = os.path.join(_DIR, "officer_codes.txt")

# --- NEW: App Config File and Defaults ---
CONFIG_FILE = os.path.join(_DIR, "elc_config.json")
DEFAULT_CLUB_NAME = "Engineering Leadership Council"
DEFAULT_OFFICER_PIN = "3132"


def _show_windows_message_box(text: str, title: str) -> None:
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(0, text, title, 0x00000010)
    except Exception:
        pass


def _write_err_log(text: str) -> None:
    try:
        with open(_ERR_LOG, "a", encoding="utf-8") as f:
            f.write("\n" + "=" * 80 + "\n")
            f.write(text)
            f.write("\n")
    except Exception:
        pass


def _excepthook(exc_type, exc_value, exc_tb):
    tb = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
    try:
        print(tb, file=sys.stderr, flush=True)
    except Exception:
        pass
    _write_err_log(tb)
    if sys.platform.startswith("win"):
        _show_windows_message_box(tb, f"{APP_TITLE} - Unhandled Error")
    sys.__excepthook__(exc_type, exc_value, exc_tb)


sys.excepthook = _excepthook


# ---------------------------------
# Tk availability (and import) check
# ---------------------------------
try:
    import tkinter as tk
    from tkinter import ttk
    import tkinter.font as tkfont
    from tkinter import simpledialog, messagebox
except Exception as _e:
    if sys.platform.startswith("win"):
        _show_windows_message_box(
            "This app needs Tkinter (Tk/Tcl). On Windows/macOS with uv’s managed Python,\n"
            "Tkinter is included. If you’re on Linux, install: python3-tk, tk, tcl.\n\n"
            f"Technical detail: {_e}",
            "Missing Tkinter",
        )
    else:
        print("Missing Tkinter/Tk/Tcl.\nError:", _e, file=sys.stderr, flush=True)
    raise


def install_tk_exception_handler(root: tk.Tk):
    def _tk_report_callback_exception(exc_type, exc_value, exc_tb):
        _excepthook(exc_type, exc_value, exc_tb)
    root.report_callback_exception = _tk_report_callback_exception


# -----------------------
# Original application
# -----------------------
from datetime import datetime
import re

# ---- SCALE ----
SCALE = 2.0

# ---- KIOSK / FULLSCREEN SETTINGS ----
# KIOSK = True <-- This is now unconditional
HIDE_CURSOR_IN_KIOSK = False

# ---- COLOR PALETTE ----
BLACK       = "#000000"
MCC_GOLD    = "#C99700"
GRAY_BG     = "#D9D9D9"
GRAY_ACTIVE = "#CFCFCF"
GRAY_PRESSED= "#C5C5C5"
GRAY_BORDER = "#A0A0A0"
SEP_COLOR   = "#333333"
GOOD_GREEN  = "#7DD97D"
BAD_RED     = "#FF5858"
WARN_YELLOW = "#FFD16A"


# Dropdown menu colors
MENU_BG            = BLACK
MENU_FG            = MCC_GOLD
MENU_ACTIVE_BG     = "#1A1A1A"
MENU_ACTIVE_FG     = MCC_GOLD
MENU_DISABLED_FG   = "#6D5A16"
MENU_BORDERWIDTH   = 0

# ---- OFFICER/ADVISOR CODES (persisted in officer_codes.txt) ----
OFFICER_CODES = {}  # code(str) -> role(str)

# ### EMAIL USERNAME + FIXED DOMAIN
STUDENT_DOMAIN = "@student.monroecc.edu"
MAX_USERNAME_LEN = 60 - len(STUDENT_DOMAIN)
USERNAME_ALLOWED_RE = re.compile(r'^[a-z0-9.\-]{0,' + str(MAX_USERNAME_LEN) + r'}$')

# ---- OFFICER MENU PIN (This is now a global placeholder, set in main()) ----
OFFICER_MENU_PIN = DEFAULT_OFFICER_PIN # Will be overwritten in main()
# --- Placeholders for live config values ---
APP_CLUB_NAME = DEFAULT_CLUB_NAME
APP_OFFICER_PIN = DEFAULT_OFFICER_PIN


def set_cursor_hidden(root: tk.Tk, hidden: bool):
    try:
        root.configure(cursor="none" if hidden else "")
    except tk.TclError:
        pass


def enter_fullscreen(root: tk.Tk):
    root.attributes("-fullscreen", True)
    set_cursor_hidden(root, HIDE_CURSOR_IN_KIOSK)


def exit_fullscreen(root: tk.Tk):
    root.attributes("-fullscreen", False)
    set_cursor_hidden(root, False)


def toggle_fullscreen(root: tk.Tk):
    fs = bool(root.attributes("-fullscreen"))
    if fs:
        exit_fullscreen(root)
    else:
        enter_fullscreen(root)


def px(value: float) -> int:
    return int(round(value * SCALE))


TEMP_ENTRIES = []


def today_str():
    return datetime.now().strftime("%B %d, %Y")


def export_txt_filename_base():
    return f"{today_str()} Attendance.txt"


def export_xlsx_filename_base():
    return f"{today_str()} Attendance.xlsx"


def get_daily_makerspace_log_path() -> str:
    """Gets path to today's log, creating it with headers if it doesn't exist."""
    filename = f"MakerSpace_Log_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    filepath = os.path.join(_DIR, filename)
    if not os.path.exists(filepath):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter as _gcl

            wb = Workbook(); ws = wb.active; ws.title = "Log"
            headers = ["First Name", "Last Name", "Email", "Sign-In Time", "Sign-Out Time", "Duration (Minutes)"]

            thin = Side(style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            header_font = Font(bold=True)
            header_fill = PatternFill("solid", fgColor="FFF2CC")

            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = header_font; cell.border = border; cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.column_dimensions[_gcl(1)].width = 20 # First
            ws.column_dimensions[_gcl(2)].width = 20 # Last
            ws.column_dimensions[_gcl(3)].width = 30 # Email
            ws.column_dimensions[_gcl(4)].width = 15 # Sign-In
            ws.column_dimensions[_gcl(5)].width = 15 # Sign-Out
            ws.column_dimensions[_gcl(6)].width = 20 # Duration

            ws.freeze_panes = ws.cell(row=2, column=1)
            wb.save(filepath)
            wb.close()
        except Exception as e:
            _write_err_log(f"Failed to create new MakerSpace log file: {e}")
            _show_windows_message_box(f"Failed to create new log file:\n{e}", "Log Creation Error")
            return None # Indicate failure
    return filepath


def unique_path(path: str) -> str:
    if not os.path.exists(path):
        return path
    root, ext = os.path.splitext(path)
    n = 2
    while True:
        candidate = f"{root} ({n}){ext}"
        if not os.path.exists(candidate):
            return candidate
        n += 1


EMAIL_ALLOWED_RE = re.compile(r'^[a-z0-9@.\-]{0,60}$')


def email_chars_ok(s: str) -> bool:
    return bool(EMAIL_ALLOWED_RE.match(s))


def is_plausible_email(s: str) -> bool:
    s = s.strip()
    if not email_chars_ok(s):
        return False
    if "@" not in s:
        return False
    user, _, domain = s.partition("@")
    if not user or not domain or "." not in domain:
        return False
    return True


def username_chars_ok(s: str) -> bool:
    return bool(USERNAME_ALLOWED_RE.match(s))


def name_len_ok(s: str) -> bool:
    return len(s) <= 50


# ---------- Officer codes persistence (TXT in same folder) ----------
def _ensure_codes_file_exists():
    if os.path.exists(OFFICER_CODES_FILE):
        return
    try:
        with open(OFFICER_CODES_FILE, "w", encoding="utf-8") as f:
            f.write("# Officer codes file - keep this file in the SAME folder as tk_sign.py\n")
            f.write("# One entry per line: CODE,Role   (comma or pipe '|' accepted)\n")
            f.write("# Example: 1234,President\n")
            f.write("# Lines starting with # are ignored.\n")
    except Exception as e:
        _write_err_log(f"Failed to create officer codes file: {e}")


def load_officer_codes() -> dict:
    codes = {}
    try:
        _ensure_codes_file_exists()
        with open(OFFICER_CODES_FILE, "r", encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if not s or s.startswith("#"):
                    continue
                parts = s.split("|", 1) if "|" in s else s.split(",", 1)
                if len(parts) != 2:
                    continue
                code, role = parts[0].strip(), parts[1].strip()
                if code.isdigit() and len(code) == 4 and role:
                    codes[code] = role
    except Exception as e:
        _write_err_log(f"Failed to load officer codes: {e}")
    return codes


def save_officer_codes(codes: dict) -> None:
    try:
        with open(OFFICER_CODES_FILE, "w", encoding="utf-8") as f:
            f.write("# Officer codes file - keep this file in the SAME folder as tk_sign.py\n")
            f.write("# CODE,Role\n")
            for code, role in sorted(codes.items()):
                f.write(f"{code},{role}\n")
    except Exception as e:
        _write_err_log(f"Failed to save officer codes: {e}")
        raise

# --- NEW: Config file helpers ---

def load_config() -> dict | None:
    """Loads config. Returns None if file doesn't exist (first run)."""
    if not os.path.exists(CONFIG_FILE):
        return None
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        _write_err_log(f"Failed to load config file: {e}")
        # File is corrupt, but exists. Treat as "not first run" but use defaults.
        return {} # Return empty dict, not None

def save_config(config_data: dict) -> None:
    """Saves config data to the JSON file."""
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config_data, f, indent=2)
    except Exception as e:
        _write_err_log(f"Failed to save config file: {e}")
        _show_windows_message_box(f"Failed to save config:\n{e}", "Config Save Error")


def main():
    # --- Root window (created early and hidden for setup) ---
    root = tk.Tk()
    root.withdraw() # Hide window during setup
    install_tk_exception_handler(root)

    # --- NEW: Config Loading (no more "one-time" setup) ---
    global APP_CLUB_NAME, APP_OFFICER_PIN, OFFICER_MENU_PIN
    
    loaded_config = load_config()
    
    if loaded_config is None: # First time run, save defaults
        app_config = {"default_club_name": DEFAULT_CLUB_NAME, "officer_pin": DEFAULT_OFFICER_PIN}
        try:
            save_config(app_config)
        except Exception as e:
             _excepthook(type(e), e, e.__traceback__)
             # Non-fatal, app will just use defaults
    else: # Config file exists
        app_config = loaded_config

    # Get final, validated config values for use in the app
    APP_CLUB_NAME = app_config.get("default_club_name", DEFAULT_CLUB_NAME)
    APP_OFFICER_PIN = app_config.get("officer_pin", DEFAULT_OFFICER_PIN)
    if not (APP_OFFICER_PIN.isdigit() and len(APP_OFFICER_PIN) == 4):
        APP_OFFICER_PIN = DEFAULT_OFFICER_PIN # Fallback if stored pin is invalid

    # Set the global PIN used by callbacks
    OFFICER_MENU_PIN = APP_OFFICER_PIN
    # --- END: Config Loading ---


    # Load officer codes at startup
    OFFICER_CODES.clear()
    OFFICER_CODES.update(load_officer_codes())

    # --- Root window (configure the one we already made) ---
    root.title(APP_TITLE)
    root.configure(bg=BLACK)

    # Layout: top bar + content
    root.grid_columnconfigure(0, weight=1)
    root.grid_rowconfigure(1, weight=1)

    root.geometry(f"{px(780)}x{px(580)}")
    root.minsize(px(640), px(520))  # keep controls visible

    try:
        root.tk.call("tk", "scaling", SCALE)
    except tk.TclError:
        pass

    # --- MODIFIED: Unconditionally enter fullscreen ---
    root.after(50, lambda: enter_fullscreen(root))

    # ESC enable/disable (locked by default)
    def _noop(_e=None): return "break"

    def set_escape_enabled(enabled: bool):
        try:
            root.unbind("<Escape>")
        except Exception:
            pass
        if enabled:
            root.bind("<Escape>", lambda e: exit_fullscreen(root))
        else:
            root.bind("<Escape>", _noop)

    set_escape_enabled(False)
    root.bind("<F11>", lambda e: toggle_fullscreen(root))

    # --- Fonts ---
    base_sz = max(9, int(round(10 * SCALE)))
    for fname in ("TkDefaultFont", "TkTextFont", "TkFixedFont", "TkMenuFont"):
        try:
            tkfont.nametofont(fname).configure(size=base_sz)
        except tk.TclError:
            pass

    title_font = tkfont.Font(
        family=tkfont.nametofont("TkDefaultFont").actual("family"),
        size=max(12, int(round(14 * SCALE))),
        weight="bold",
    )
    hint_font = tkfont.Font(
        family=tkfont.nametofont("TkDefaultFont").actual("family"),
        size=max(8, int(round((base_sz - 2)))),
    )

    # --- THEME / STYLES ---
    style = ttk.Style()
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass

    style.configure(".", background=BLACK, foreground=MCC_GOLD)
    style.configure("TFrame", background=BLACK)
    style.configure("TLabel", background=BLACK, foreground=MCC_GOLD)
    style.configure("TSeparator", background=SEP_COLOR)

    style.configure("Good.TLabel", background=BLACK, foreground=GOOD_GREEN)
    style.configure("Hint.TLabel", background=BLACK, foreground=MCC_GOLD)

    style.configure(
        "CTM.TEntry",
        fieldbackground=GRAY_BG,
        foreground="#111111",
        insertcolor="#111111",
        bordercolor=GRAY_BORDER,
        lightcolor=GRAY_BORDER,
        darkcolor=GRAY_BORDER,
        padding=(px(6), px(4)),
        relief="flat",
    )
    style.configure(
        "CTM.TButton",
        background=GRAY_BG,
        foreground="#111111",
        bordercolor=GRAY_BORDER,
        padding=(px(16), px(10)),
        relief="flat",
    )
    style.map(
        "CTM.TButton",
        background=[("active", GRAY_ACTIVE), ("pressed", GRAY_PRESSED)],
        relief=[("pressed", "flat"), ("!pressed", "flat")],
        foreground=[("disabled", "#666666")]
    )
    
    # --- NEW: Small button style for Officer Manager ---
    style.configure(
        "Small.CTM.TButton",
        background=GRAY_BG,
        foreground="#111111",
        bordercolor=GRAY_BORDER,
        padding=(px(5), px(3)), # <-- Reduced padding
        relief="flat",
    )
    style.map(
        "Small.CTM.TButton",
        background=[("active", GRAY_ACTIVE), ("pressed", GRAY_PRESSED)],
        relief=[("pressed", "flat"), ("!pressed", "flat")],
        foreground=[("disabled", "#666666")]
    )
    # --- End new style ---

    style.configure("CTM.TCheckbutton", background=BLACK, foreground=MCC_GOLD)

    # Treeview styling
    style.configure(
        "ELC.Treeview",
        background=BLACK, fieldbackground=BLACK, foreground=MCC_GOLD,
        rowheight=px(28)
    )
    style.configure("ELC.Treeview.Heading", background=BLACK, foreground=MCC_GOLD)

    # Tk classic menu theming
    root.option_add('*Menu.background', MENU_BG)
    root.option_add('*Menu.foreground', MENU_FG)
    root.option_add('*Menu.activeBackground', MENU_ACTIVE_BG)
    root.option_add('*Menu.activeForeground', MENU_ACTIVE_FG)
    root.option_add('*Menu.disabledForeground', MENU_DISABLED_FG)
    root.option_add('*Menu.borderWidth', MENU_BORDERWIDTH)

    # --- State Vars (shared across modes) ---
    club_org_var       = tk.StringVar(value=APP_CLUB_NAME) # <-- MODIFIED
    meeting_topic_var  = tk.StringVar(value="")
    interest_input_var = tk.StringVar(value="")

    first_var = tk.StringVar()
    last_var  = tk.StringVar()
    username_var = tk.StringVar()

    # --- State Vars (Event Mode) ---
    queued_num_var = tk.StringVar(value="0")
    status_var     = tk.StringVar(value="") # Event mode status
    officer_code_var = tk.StringVar(value="")
    officer_list_var = tk.StringVar(value="")
    officer_count_var = tk.StringVar(value="0")
    present_officers = set()

    status_lbl  = None # Event mode status label
    first_entry = None # Event mode first entry widget
    
    # --- NEW: Reference to the current main UI frame ---
    current_main_frame = None

    interest_labels = []
    interest_vars   = []

    # --- Validation ---
    vcmd_username = (root.register(lambda v: v == "" or username_chars_ok(v)), "%P")
    vcmd_name     = (root.register(lambda v: len(v) <= 50), "%P")
    vcmd_officer  = (root.register(lambda v: (v == "") or (v.isdigit() and len(v) <= 4)), "%P")

    # --- Callbacks (Event Mode) ---
    def set_status(msg: str, color: str):
        nonlocal status_lbl
        status_var.set(msg)
        if status_lbl is not None:
            status_lbl.configure(foreground=color)

    def update_queue_ui():
        queued_num_var.set(str(len(TEMP_ENTRIES)))

    def submit_callback():
        nonlocal first_entry
        first = first_var.get().strip()
        last  = last_var.get().strip()
        username = username_var.get().strip().lower()
        username_var.set(username)

        email = f"{username}{STUDENT_DOMAIN}"

        if not first or not last or not username:
            set_status("Please fill all fields (first, last, and email username).", BAD_RED)
            return
        if not is_plausible_email(email):
            set_status("Email looks invalid. Please use letters, numbers, dot or dash.", BAD_RED)
            return

        interests_bits = [("1" if v.get() else "0") for v in interest_vars]
        TEMP_ENTRIES.append([first, last, email] + interests_bits)

        set_status(f"Submitted: {first}, {last}, {email}", GOOD_GREEN)

        first_var.set(""); last_var.set(""); username_var.set("")
        for v in interest_vars: v.set(False)
        if first_entry is not None: first_entry.focus_set()
        update_queue_ui()

    def record_officer_code():
        code = officer_code_var.get().strip()
        if not code or len(code) != 4 or not code.isdigit():
            set_status("Invalid code (must be 4 digits).", WARN_YELLOW); return
        role = OFFICER_CODES.get(code)
        if not role:
            set_status("Code not recognized.", WARN_YELLOW); officer_code_var.set(""); return
        if role in present_officers:
            set_status(f"{role} already recorded.", WARN_YELLOW); officer_code_var.set(""); return
        present_officers.add(role)
        officer_count_var.set(str(len(present_officers)))
        officer_code_var.set("")
        set_status("Officer recorded", MCC_GOLD)

    def export_callback():
        topic = meeting_topic_var.get().strip()
        club  = club_org_var.get().strip() or "Engineering Leadership Council"
        if not TEMP_ENTRIES:
            set_status("No submissions to export.", WARN_YELLOW); return

        def clean_header(s: str, idx: int) -> str:
            import re as _re
            s = (s or "").replace("\r", " ").replace("\n", " ").replace("\t", " ")
            s = _re.sub(r"\s+", " ", s).strip()
            if not s: s = f"Interest {idx+1}"
            if s.startswith("="): s = "'" + s
            return s[:60]

        safe_interest_headers = [clean_header(lbl, i) for i, lbl in enumerate(interest_labels)]
        base_headers = ["First Name", "Last Name", "Email"]
        headers = base_headers + (safe_interest_headers if safe_interest_headers else ["Interests"])

        rows = []
        for record in TEMP_ENTRIES:
            first, last, email, *bits = record
            if safe_interest_headers:
                checks = [("✓" if (i < len(bits) and bits[i] == "1") else "") for i in range(len(safe_interest_headers))]
                rows.append([first, last, email] + checks)
            else:
                selected = [lbl for lbl, bit in zip(interest_labels, bits) if bit == "1"]
                rows.append([first, last, email, ", ".join(selected)])

        xlsx_name = unique_path(export_xlsx_filename_base())

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter as _gcl

            wb = Workbook(); ws = wb.active; ws.title = "Attendance"

            thin = Side(style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            header_font = Font(bold=True)
            title_font  = Font(bold=True, size=14)

            INTEREST_COLORS = ["E6F7FF","E8F5E9","FFF3E0","F3E5F5","FFFDE7","E1F5FE","FCE4EC","E0F2F1"]

            ws["A1"] = today_str(); ws["A1"].font = title_font
            ws["A3"] = club;        ws["A3"].font = title_font
            ws["A5"] = f"Meeting: {topic}" if topic else "Meeting:"
            ws["A6"] = f"Members of the board present: {', '.join(sorted(present_officers))}"

            start_row = 8
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=c, value=h)
                cell.font = header_font; cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if c <= 3 or not safe_interest_headers:
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                else:
                    color = INTEREST_COLORS[(c - 4) % len(INTEREST_COLORS)]
                    cell.fill = PatternFill("solid", fgColor=color)

            for r_i, row in enumerate(rows, start=start_row + 1):
                for c_i, val in enumerate(row, start=1):
                    cell = ws.cell(row=r_i, column=c_i, value=val)
                    cell.border = border
                    if c_i <= 3 or not safe_interest_headers:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        if val == "✓":
                            color = INTEREST_COLORS[(c_i - 4) % len(INTEREST_COLORS)]
                            cell.fill = PatternFill("solid", fgColor=color)

            ws.freeze_panes = ws.cell(row=start_row + 1, column=4 if safe_interest_headers else 1)

            last_col = len(headers); last_row = start_row + len(rows)
            ws.auto_filter.ref = f"A{start_row}:{_gcl(last_col)}{last_row}"

            col_widths = [len(h) for h in headers]
            for row in rows:
                for i, val in enumerate(row):
                    col_widths[i] = max(col_widths[i], len(str(val)) if val is not None else 0)
            for i, w in enumerate(col_widths, start=1):
                is_interest_col = (i > 3) and bool(safe_interest_headers)
                width = 10 if is_interest_col else min(60, max(12, w + 2))
                ws.column_dimensions[_gcl(i)].width = width

            wb.save(xlsx_name)
            set_status(f"Exported {len(TEMP_ENTRIES)} entrie(s) to Excel: {os.path.abspath(xlsx_name)}", GOOD_GREEN)
            TEMP_ENTRIES.clear(); update_queue_ui()

        except ImportError:
            import csv
            csv_name = unique_path(export_txt_filename_base().replace(".txt", ".csv"))
            try:
                with open(csv_name, "w", newline="", encoding="utf-8") as f:
                    w = csv.writer(f)
                    w.writerow([today_str()]); w.writerow([]); w.writerow([club]); w.writerow([])
                    w.writerow([f"Meeting: {topic}"] if topic else ["Meeting:"])
                    w.writerow([f"Members of the board present: {', '.join(sorted(present_officers))}"])
                    w.writerow([]); w.writerow(headers)
                    for row in rows:
                        w.writerow([("Y" if (cell == "✓") else cell) for cell in row])
                set_status(f"openpyxl not installed. Exported CSV instead: {os.path.abspath(csv_name)}", WARN_YELLOW)
                TEMP_ENTRIES.clear(); update_queue_ui()
            except Exception as e:
                set_status(f"CSV export failed: {e}", BAD_RED)
        except Exception as e:
            set_status(f"Excel export failed: {e}", BAD_RED)

    # -----------------------
    # Settings menu (with Officer Codes manager)
    # -----------------------
    officer_unlocked = False  # locked by default

    def _valid_pin(pin: str) -> bool:
        return pin.isdigit() and len(pin) == 4

    def _ask_officer_pin():
        return simpledialog.askstring("Officer Access", "Enter 4-digit passcode:", show="*", parent=root)

    def _make_menu(parent):
        return tk.Menu(
            parent, tearoff=False,
            bg=MENU_BG, fg=MENU_FG,
            activebackground=MENU_ACTIVE_BG, activeforeground=MENU_ACTIVE_FG,
            disabledforeground=MENU_DISABLED_FG, borderwidth=MENU_BORDERWIDTH
        )

    # Top bar with "Settings"
    topbar = ttk.Frame(root, style="TFrame", padding=(px(12), px(6)))
    topbar.grid(row=0, column=0, sticky="ew")
    topbar.grid_columnconfigure(0, weight=1)

    style.configure("OfficerBar.TMenubutton", background=BLACK, foreground=MCC_GOLD,
                    relief="flat", padding=(px(10), px(4)))
    style.map("OfficerBar.TMenubutton",
              background=[("active", "#0d0d0d")],
              foreground=[("disabled", MENU_DISABLED_FG)])

    settings_btn = ttk.Menubutton(topbar, text="Settings", style="OfficerBar.TMenubutton")
    settings_btn.grid(row=0, column=0, sticky="w")

    settings_menu = _make_menu(settings_btn)
    settings_btn["menu"] = settings_menu

    # 0) Toggle lock/unlock (label changes at runtime)
    settings_menu.add_command(label="Unlock Settings…", command=lambda: on_toggle_lock())
    idx_toggle = settings_menu.index("end")

    settings_menu.add_separator()

    # Officer Codes submenu
    officer_codes_menu = _make_menu(settings_menu)
    settings_menu.add_cascade(label="Officer Codes", menu=officer_codes_menu)
    idx_codes_cascade = settings_menu.index("end")

    def open_officer_codes_manager():
        mgr = tk.Toplevel(root)
        mgr.title("Officer Codes Manager")
        mgr.configure(bg=BLACK)
        mgr.transient(root)
        mgr.grab_set()

        # --- MODIFIED: Make Officer Manager fullscreen ---
        mgr.after(50, lambda: enter_fullscreen(mgr))
        mgr.bind("<F11>", lambda e: toggle_fullscreen(mgr))
        mgr.bind("<Escape>", lambda e: exit_fullscreen(mgr))
        # --- End modification ---

        body = ttk.Frame(mgr, style="TFrame", padding=(px(12), px(12)))
        body.pack(fill="both", expand=True)

        table_wrap = ttk.Frame(body, style="TFrame")
        table_wrap.pack(fill="both", expand=True)

        columns = ("code", "role")
        tree = ttk.Treeview(
            table_wrap, columns=columns, show="headings", height=10, style="ELC.Treeview"
        )
        tree.grid(row=0, column=0, sticky="nsew")
        table_wrap.grid_columnconfigure(0, weight=1)
        table_wrap.grid_rowconfigure(0, weight=1)

        sb = ttk.Scrollbar(table_wrap, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        sb.grid(row=0, column=1, sticky="ns")

        tree.heading("code", text="Code")
        tree.heading("role", text="Role")
        tree.column("code", width=px(120), minwidth=px(100), stretch=False, anchor="center")
        tree.column("role", width=px(360), minwidth=px(260), stretch=True, anchor="w")

        form = ttk.Frame(body, style="TFrame")
        form.pack(fill="x", pady=(px(12), px(4)))

        ttk.Label(form, text="Code (4 digits):", style="TLabel").grid(row=0, column=0, sticky="w")
        code_var = tk.StringVar()
        code_entry = ttk.Entry(form, textvariable=code_var, width=8, style="CTM.TEntry",
                               validate="key", validatecommand=vcmd_officer)
        code_entry.grid(row=0, column=1, sticky="w", padx=(px(6), px(20)))

        ttk.Label(form, text="Role:", style="TLabel").grid(row=0, column=2, sticky="w")
        role_var = tk.StringVar()
        role_entry = ttk.Entry(form, textvariable=role_var, width=30, style="CTM.TEntry",
                               validate="key", validatecommand=(root.register(lambda v: len(v) <= 50), "%P"))
        role_entry.grid(row=0, column=3, sticky="ew", padx=(px(6), 0))
        form.grid_columnconfigure(3, weight=1)

        btns = ttk.Frame(body, style="TFrame")
        btns.pack(fill="x", pady=(px(6), 0))

        def refresh_table():
            for i in tree.get_children():
                tree.delete(i)
            for code, role in sorted(OFFICER_CODES.items()):
                tree.insert("", "end", values=(code, role))

        def on_add_update():
            code = code_var.get().strip()
            role = role_var.get().strip()
            if not (code.isdigit() and len(code) == 4):
                messagebox.showerror(APP_TITLE, "Code must be exactly 4 digits.", parent=mgr); return
            if not role:
                messagebox.showerror(APP_TITLE, "Role cannot be empty.", parent=mgr); return
            OFFICER_CODES[code] = role
            refresh_table()
            code_var.set(""); role_var.set("")

        def on_delete():
            sel = tree.selection()
            if not sel:
                return
            for item in sel:
                code = tree.item(item, "values")[0]
                OFFICER_CODES.pop(code, None)
            refresh_table()

        def on_reload():
            loaded = load_officer_codes()
            if loaded is not None:
                OFFICER_CODES.clear()
                OFFICER_CODES.update(loaded)
                refresh_table()
            else:
                messagebox.showwarning(APP_TITLE, "No codes loaded (file empty or invalid).", parent=mgr)

        def on_save():
            try:
                save_officer_codes(OFFICER_CODES)
                messagebox.showinfo(APP_TITLE, "Officer codes saved.", parent=mgr)
            except Exception as e:
                messagebox.showerror(APP_TITLE, f"Save failed: {e}", parent=mgr)

        # --- MODIFIED: Use "Small.CTM.TButton" style ---
        ttk.Button(btns, text="Add / Update",      style="Small.CTM.TButton", width=16, command=on_add_update).pack(side="left")
        ttk.Button(btns, text="Delete Selected",   style="Small.CTM.TButton", width=16, command=on_delete).pack(side="left", padx=(px(8), 0))
        ttk.Button(btns, text="Reload from File", style="Small.CTM.TButton", width=16, command=on_reload).pack(side="left", padx=(px(8), 0))
        ttk.Button(btns, text="Save to File",      style="Small.CTM.TButton", width=16, command=on_save).pack(side="left", padx=(px(8), 0))
        ttk.Button(btns, text="Close",             style="Small.CTM.TButton", width=10, command=mgr.destroy).pack(side="right")
        # --- End modification ---

        def on_row_select(event=None):
            sel = tree.selection()
            if not sel:
                return
            code, role = tree.item(sel[0], "values")
            code_var.set(code); role_var.set(role)

        tree.bind("<<TreeviewSelect>>", on_row_select)
        refresh_table()
        code_entry.focus_set()

        mgr.update_idletasks()
        # These are now just fallbacks if fullscreen fails
        req_w = max(mgr.winfo_reqwidth(), px(560))
        req_h = max(mgr.winfo_reqheight(), px(520))
        mgr.minsize(req_w, req_h)
        mgr.geometry(f"{req_w}x{req_h}")

    officer_codes_menu.add_command(label="Open Manager…", command=open_officer_codes_manager)
    idx_codes_open = officer_codes_menu.index("end")

    # Remaining Settings actions
    settings_menu.add_command(label="Export (Event Mode)", command=export_callback, state="disabled")
    idx_export = settings_menu.index("end")
    settings_menu.add_command(label="Toggle Fullscreen (F11)", command=lambda: toggle_fullscreen(root), state="disabled")
    idx_full = settings_menu.index("end")
    settings_menu.add_command(
        label="Clear Queue (Event Mode)",
        command=lambda: (TEMP_ENTRIES.clear(), update_queue_ui(), set_status("Queue cleared.", MCC_GOLD))
                      if TEMP_ENTRIES else messagebox.showinfo(APP_TITLE, "Queue is already empty.", parent=root),
        state="disabled"
    )
    idx_clear = settings_menu.index("end")
    
    # --- NEW: Go Back functionality ---
    def go_back_to_mode_select():
        nonlocal current_main_frame
        
        # 1. Hide the current active frame (if one exists)
        if current_main_frame:
            current_main_frame.grid_forget()
            current_main_frame = None # Clear the reference
        
        # 2. Hide the event prompt frame (in case we are there)
        prompt.grid_forget()
        
        # 3. Show the main mode selection frame
        mode_frm.grid(row=1, column=0, sticky="nsew")
        event_btn.focus_set() # Focus the first button
        
        # 4. Reset app state
        root.title(APP_TITLE)
        
        # Clear shared vars
        first_var.set("")
        last_var.set("")
        username_var.set("")
        
        # Reset event mode state
        TEMP_ENTRIES.clear()
        present_officers.clear()
        update_queue_ui() # Resets queue_num_var
        officer_count_var.set("0")
        status_var.set("")
        
        # Reset prompt vars
        club_org_var.set(APP_CLUB_NAME)
        meeting_topic_var.set("")
        interest_input_var.set("")
        
        # (MakerSpace status var is local to its build function, so it's recreated next time)
        
        # 5. Show a message
        messagebox.showinfo(APP_TITLE, "Returned to main menu. App state has been reset.", parent=root)

    settings_menu.add_separator()
    settings_menu.add_command(label="< Back to Main Menu", command=go_back_to_mode_select, state="disabled")
    idx_back_to_menu = settings_menu.index("end")
    # --- END: Go Back functionality ---

    def _refresh_settings_menu():
        settings_menu.entryconfig(idx_toggle, label=("Lock Settings" if officer_unlocked else "Unlock Settings…"))
        state = "normal" if officer_unlocked else "disabled"
        settings_menu.entryconfig(idx_codes_cascade, state=state)
        officer_codes_menu.entryconfig(idx_codes_open, state=state)
        settings_menu.entryconfig(idx_export, state=state)
        settings_menu.entryconfig(idx_full,   state=state)
        settings_menu.entryconfig(idx_clear,  state=state)
        settings_menu.entryconfig(idx_back_to_menu, state=state) # <-- NEW

    def on_toggle_lock():
        nonlocal officer_unlocked
        if officer_unlocked:
            officer_unlocked = False
            set_escape_enabled(False)
            _refresh_settings_menu()
            messagebox.showinfo("Officer Access", "Settings locked.", parent=root)
        else:
            pin = _ask_officer_pin()
            if pin is None:
                return
            if not _valid_pin(pin):
                messagebox.showerror("Officer Access", "PIN must be exactly 4 digits.", parent=root); return
            if pin != OFFICER_MENU_PIN: # <-- MODIFIED (uses global var)
                messagebox.showerror("Officer Access", "Incorrect passcode.", parent=root); return
            officer_unlocked = True
            set_escape_enabled(True)
            _refresh_settings_menu()
            messagebox.showinfo("Officer Access", "Settings unlocked.", parent=root)

    _refresh_settings_menu()  # initialize locked

    # --- MakerSpace UI (split left/right)
    def build_makerspace_ui():
        nonlocal first_var, last_var, username_var, vcmd_name, vcmd_username, root, title_font, hint_font
        nonlocal current_main_frame # <-- NEW

        root.title("MakerSpace Sign-In")

        # Local state for this UI
        ms_status_var = tk.StringVar(value="Welcome! Enter your info to sign in.")
        ms_status_lbl = None
        ms_first_entry = None

        frm = ttk.Frame(root, padding=(px(24), px(24)), style="TFrame")
        frm.grid(row=1, column=0, sticky="nsew")
        frm.grid_columnconfigure(0, weight=1)
        frm.grid_columnconfigure(1, weight=1)
        frm.grid_rowconfigure(1, weight=1)

        ttk.Label(frm, text="MakerSpace / Room Sign-In", font=title_font, style="TLabel")\
            .grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, px(12)))

        # LEFT
        left = ttk.Frame(frm, style="TFrame")
        left.grid(row=1, column=0, sticky="nsew", padx=(0, px(12)))
        left.grid_columnconfigure(0, weight=1)

        def ms_set_status(msg: str, color: str):
            nonlocal ms_status_lbl
            ms_status_var.set(msg)
            if ms_status_lbl is not None:
                ms_status_lbl.configure(foreground=color)

        # First / Last / Email username
        ttk.Label(left, text="First name", style="TLabel").grid(row=1, column=0, sticky="w", pady=(px(6), px(4)))
        ms_first_entry = ttk.Entry(left, textvariable=first_var, width=28, style="CTM.TEntry",
                                   validate="key", validatecommand=vcmd_name)
        ms_first_entry.grid(row=2, column=0, sticky="ew", pady=(0, px(10)))
        ms_first_entry.focus_set()

        ttk.Label(left, text="Last name", style="TLabel").grid(row=3, column=0, sticky="w", pady=(px(6), px(4)))
        last_entry = ttk.Entry(left, textvariable=last_var, width=28, style="CTM.TEntry",
                               validate="key", validatecommand=vcmd_name)
        last_entry.grid(row=4, column=0, sticky="ew", pady=(0, px(10)))

        ttk.Label(left, text="Email username", style="TLabel").grid(row=5, column=0, sticky="w", pady=(px(6), px(4)))
        email_wrap = ttk.Frame(left, style="TFrame")
        email_wrap.grid(row=6, column=0, sticky="w", pady=(0, 0))
        user_entry = ttk.Entry(
            email_wrap, textvariable=username_var, width=20, style="CTM.TEntry",
            validate="key", validatecommand=vcmd_username
        )
        user_entry.grid(row=0, column=0, sticky="w")

        # --- Callbacks for this mode ---

        def clear_fields():
            nonlocal ms_first_entry
            first_var.set(""); last_var.set(""); username_var.set("")
            if ms_first_entry: ms_first_entry.focus_set()

        def get_validated_input():
            first = first_var.get().strip()
            last  = last_var.get().strip()
            username = username_var.get().strip().lower()
            username_var.set(username)
            email = f"{username}{STUDENT_DOMAIN}"

            if not first or not last or not username:
                ms_set_status("Please fill all fields (first, last, and email username).", BAD_RED)
                return None
            if not is_plausible_email(email):
                ms_set_status("Email looks invalid. Please use letters, numbers, dot or dash.", BAD_RED)
                return None
            return (first, last, email)

        def makerspace_sign_in_callback():
            inputs = get_validated_input()
            if not inputs: return
            first, last, email = inputs
            now = datetime.now()

            try:
                from openpyxl import load_workbook
            except ImportError:
                ms_set_status("ERROR: openpyxl not found. Cannot write to log.", BAD_RED)
                return

            filepath = get_daily_makerspace_log_path()
            if not filepath:
                ms_set_status("ERROR: Could not create or access log file.", BAD_RED); return

            try:
                wb = load_workbook(filepath)
                ws = wb.active
                # Check for existing open session
                for row in range(ws.max_row, 1, -1): # Iterate backwards from end
                    if ws.cell(row, 3).value == email and ws.cell(row, 5).value is None:
                        ms_set_status(f"You are already signed in (at {ws.cell(row, 4).value}). Sign out first.", WARN_YELLOW)
                        wb.close(); return

                # Add new sign-in row
                new_row = [first, last, email, now.strftime("%H:%M:%S"), None, None]
                ws.append(new_row)
                wb.save(filepath)
                wb.close()
                ms_set_status(f"Signed In: {first} {last}", GOOD_GREEN)
                clear_fields()

            except PermissionError:
                ms_set_status("ERROR: Log file is open in Excel. Please close it and try again.", BAD_RED)
            except Exception as e:
                _write_err_log(f"MakerSpace Sign-In Error: {e}")
                ms_set_status(f"An error occurred: {e}", BAD_RED)

        def makerspace_sign_out_callback():
            inputs = get_validated_input()
            if not inputs: return
            first, last, email = inputs # 'first' and 'last' are just for validation, 'email' is key
            now = datetime.now()

            try:
                from openpyxl import load_workbook
            except ImportError:
                ms_set_status("ERROR: openpyxl not found. Cannot write to log.", BAD_RED)
                return

            filepath = get_daily_makerspace_log_path()
            if not filepath or not os.path.exists(filepath):
                ms_set_status("ERROR: No log file found. Cannot sign out.", BAD_RED); return

            try:
                wb = load_workbook(filepath)
                ws = wb.active
                found_row_idx = -1
                # Find the user's open session
                for row_idx in range(ws.max_row, 1, -1): # Iterate backwards
                    if ws.cell(row_idx, 3).value == email and ws.cell(row_idx, 5).value is None:
                        found_row_idx = row_idx
                        break

                if found_row_idx == -1:
                    ms_set_status("Could not find an open sign-in for this email.", WARN_YELLOW)
                    wb.close(); return

                # Found it, let's update
                sign_in_str = ws.cell(found_row_idx, 4).value
                sign_in_time = datetime.strptime(sign_in_str, "%H:%M:%S")
                # Combine sign-in time with today's date for correct duration calculation
                sign_in_dt = now.replace(hour=sign_in_time.hour, minute=sign_in_time.minute, second=sign_in_time.second, microsecond=0)
                
                # Handle sign-out past midnight
                if now < sign_in_dt:
                    # Assume they signed in yesterday, this is complex.
                    # Simple approach: just log sign-out time, skip duration.
                    duration_minutes = "N/A (Past Midnight)"
                else:
                    duration = now - sign_in_dt
                    duration_minutes = round(duration.total_seconds() / 60, 1)

                ws.cell(found_row_idx, 5, value=now.strftime("%H:%M:%S"))
                ws.cell(found_row_idx, 6, value=duration_minutes)

                wb.save(filepath)
                wb.close()
                ms_set_status(f"Signed Out: {first} {last}. Duration: {duration_minutes} min.", GOOD_GREEN)
                clear_fields()

            except PermissionError:
                ms_set_status("ERROR: Log file is open in Excel. Please close it and try again.", BAD_RED)
            except Exception as e:
                _write_err_log(f"MakerSpace Sign-Out Error: {e}")
                ms_set_status(f"An error occurred: {e}", BAD_RED)


        # --- Buttons ---
        btn_frame = ttk.Frame(left, style="TFrame")
        btn_frame.grid(row=7, column=0, sticky="w", pady=(px(8), px(0)))

        sign_in_btn = ttk.Button(
            btn_frame, text="Sign IN", style="CTM.TButton",
            command=makerspace_sign_in_callback, width=14
        )
        sign_in_btn.pack(side="left", fill="x", padx=(0, px(8)))

        sign_out_btn = ttk.Button(
            btn_frame, text="Sign OUT", style="CTM.TButton",
            command=makerspace_sign_out_callback, width=14
        )
        sign_out_btn.pack(side="left", fill="x")

        # Enter key in email box triggers sign-in
        user_entry.bind("<Return>", lambda e: makerspace_sign_in_callback())

        ms_status_lbl = ttk.Label(left, textvariable=ms_status_var, style="Good.TLabel", justify="left")
        ms_status_lbl.grid(row=8, column=0, sticky="w", pady=(px(8), 0))

        def _update_status_wrap(event=None):
            try:
                width = left.winfo_width()
                ms_status_lbl.configure(wraplength=max(px(260), width - px(24)))
            except Exception: pass
        left.bind("<Configure>", _update_status_wrap)
        _update_status_wrap()

        # RIGHT
        right = ttk.Frame(frm, style="TFrame")
        right.grid(row=1, column=1, sticky="nsew", padx=(px(12), 0))
        right.grid_columnconfigure(0, weight=1)

        instructions_text = (
            "How to use this terminal:\n\n"
            "• Signing IN:\n"
            "  1. Fill your First, Last, and Email Username.\n"
            "  2. Click 'Sign IN'.\n\n"
            "• Signing OUT:\n"
            "  1. Fill the same info (especially Email).\n"
            "  2. Click 'Sign OUT'.\n\n"
            "Your time in the room will be logged automatically."
        )
        instructions = ttk.Label(right, text=instructions_text, style="Hint.TLabel", font=hint_font, justify="left")
        instructions.grid(row=0, column=0, sticky="nw", padx=(px(2), 0), pady=(px(20), 0))

        def _update_instr_wrap(event=None):
            try:
                width = right.winfo_width()
                instructions.configure(wraplength=max(px(280), width - px(24)))
            except Exception: pass
        right.bind("<Configure>", _update_instr_wrap)
        _update_instr_wrap()
        
        current_main_frame = frm # <-- NEW

    # --- Sign-in UI (split left/right) + interest cap + instructions
    def build_signin_ui():
        nonlocal status_lbl, first_entry, interest_labels, interest_vars
        nonlocal current_main_frame # <-- NEW

        frm = ttk.Frame(root, padding=(px(24), px(24)), style="TFrame")
        frm.grid(row=1, column=0, sticky="nsew")
        frm.grid_columnconfigure(0, weight=1)
        frm.grid_columnconfigure(1, weight=1)
        frm.grid_rowconfigure(1, weight=1)

        ttk.Label(frm, text=f"{club_org_var.get()} Sign-In", font=title_font, style="TLabel")\
            .grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, px(12)))

        # LEFT
        left = ttk.Frame(frm, style="TFrame")
        left.grid(row=1, column=0, sticky="nsew", padx=(0, px(12)))
        left.grid_columnconfigure(0, weight=1)

        # Compact header row: Board present + Queued on one line
        present_row = ttk.Frame(left, style="TFrame")
        present_row.grid(row=0, column=0, sticky="ew", pady=(0, px(6)))
        present_row.grid_columnconfigure(1, weight=1)

        ttk.Label(present_row, text="Board present:", style="TLabel")\
           .grid(row=0, column=0, sticky="w")

        officers_lbl = ttk.Label(present_row, textvariable=officer_count_var, style="Good.TLabel")
        officers_lbl.grid(row=0, column=1, sticky="w", padx=(px(6), 0))

        ttk.Label(present_row, text="Queued:", style="TLabel")\
           .grid(row=0, column=2, sticky="e", padx=(px(12), px(4)))
        ttk.Label(present_row, textvariable=queued_num_var, style="Good.TLabel")\
           .grid(row=0, column=3, sticky="e")

        # First / Last / Email username
        ttk.Label(left, text="First name", style="TLabel").grid(row=1, column=0, sticky="w", pady=(px(6), px(4)))
        first_entry = ttk.Entry(left, textvariable=first_var, width=28, style="CTM.TEntry",
                                validate="key", validatecommand=vcmd_name)
        first_entry.grid(row=2, column=0, sticky="ew", pady=(0, px(10)))

        ttk.Label(left, text="Last name", style="TLabel").grid(row=3, column=0, sticky="w", pady=(px(6), px(4)))
        last_entry = ttk.Entry(left, textvariable=last_var, width=28, style="CTM.TEntry",
                               validate="key", validatecommand=vcmd_name)
        last_entry.grid(row=4, column=0, sticky="ew", pady=(0, px(10)))

        ttk.Label(left, text="Email username", style="TLabel").grid(row=5, column=0, sticky="w", pady=(px(6), px(4)))

        # Username on a single line (no domain text shown)
        email_wrap = ttk.Frame(left, style="TFrame")
        email_wrap.grid(row=6, column=0, sticky="w", pady=(0, 0))
        user_entry = ttk.Entry(
            email_wrap, textvariable=username_var, width=20, style="CTM.TEntry",
            validate="key", validatecommand=vcmd_username
        )
        user_entry.grid(row=0, column=0, sticky="w")

        # Record button directly UNDERNEATH the username box
        submit = ttk.Button(
            left, text="Record Entry", style="CTM.TButton",
            command=submit_callback, width=14
        )
        submit.grid(row=7, column=0, sticky="w", pady=(px(8), px(0)))

        # Enter key submits when typing the email username
        user_entry.bind("<Return>", lambda e: submit_callback())

        status_lbl = ttk.Label(left, textvariable=status_var, style="Good.TLabel", justify="left")
        status_lbl.grid(row=8, column=0, sticky="w", pady=(px(8), 0))

        def _update_status_wrap(event=None):
            try:
                width = left.winfo_width()
                status_lbl.configure(wraplength=max(px(260), width - px(24)))
            except Exception:
                pass
        left.bind("<Configure>", _update_status_wrap)
        _update_status_wrap()

        # RIGHT
        right = ttk.Frame(frm, style="TFrame")
        right.grid(row=1, column=1, sticky="nsew", padx=(px(12), 0))
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(1, weight=0)

        officer_inline = ttk.Frame(right, style="TFrame")
        officer_inline.grid(row=0, column=0, sticky="e", pady=(0, px(8)))
        ttk.Label(officer_inline, text="Officer Code:", style="TLabel").grid(row=0, column=0, sticky="e", padx=(0, px(6)))
        off_entry = ttk.Entry(officer_inline, textvariable=officer_code_var, width=6, style="CTM.TEntry",
                               validate="key", validatecommand=vcmd_officer)
        off_entry.grid(row=0, column=1, sticky="w")
        ttk.Button(officer_inline, text="✓", style="CTM.TButton", command=record_officer_code)\
            .grid(row=0, column=2, sticky="w", padx=(px(8), 0))

        # Interests (gold box)
        outer = None
        inner = None
        interests_row = 1

        if interest_labels:
            outer = tk.Frame(right, bg=MCC_GOLD, bd=0, highlightthickness=0)
            outer.grid(row=interests_row, column=0, sticky="nw", pady=(px(10), px(6)))
            outer.grid_propagate(False)

            inner = tk.Frame(outer, bg=BLACK)
            inner.pack(padx=px(2), pady=px(2), fill="both", expand=True)

            ttk.Label(inner, text="What are your interests?", style="TLabel")\
                .grid(row=0, column=0, sticky="w", pady=(px(10), px(6)), padx=(px(10), 0))

            grid = ttk.Frame(inner, style="TFrame")
            grid.grid(row=1, column=0, sticky="nw", padx=(px(10), px(10)), pady=(0, px(10)))

            interest_vars.clear()
            for i, lbl in enumerate(interest_labels):
                var = tk.BooleanVar(value=False)
                interest_vars.append(var)
                cb = ttk.Checkbutton(grid, text=lbl, variable=var, style="CTM.TCheckbutton")
                cb.grid(row=i // 3, column=i % 3, sticky="w", padx=(0, px(16)), pady=(0, px(8)))
        else:
            ttk.Frame(right, style="TFrame").grid(row=interests_row, column=0, sticky="w", pady=(px(10), px(6)))

        instructions_text = (
            "How to fill the form:\n"
            "• First/Last Name in the proper fields.\n"
            "• Email username. (WITHOUT the @student.monroecc.edu part.)\n"
            "• Interests: check all that apply (optional).\n"
            "• Click “Record Entry” to submit."
        )
        instructions = ttk.Label(right, text=instructions_text, style="Hint.TLabel", font=hint_font, justify="left")
        instructions.grid(row=interests_row + 1, column=0, sticky="nw", padx=(px(2), 0), pady=(px(4), 0))

        def _update_instr_wrap(event=None):
            try:
                width = right.winfo_width()
                instructions.configure(wraplength=max(px(280), width - px(24)))
            except Exception:
                pass
        right.bind("<Configure>", _update_instr_wrap)
        _update_instr_wrap()

        # Cap interests height
        def _cap_interests_height(event=None):
            try:
                if outer is None:
                    return
                frm.update_idletasks()
                first_bottom = first_entry.winfo_y() + first_entry.winfo_height()
                outer_top_in_root = outer.winfo_rooty()
                frm_top_in_root   = frm.winfo_rooty()
                outer_top = outer_top_in_root - frm_top_in_root
                max_h = max(px(40), first_bottom - outer_top)
                outer.configure(height=max_h, width=max(px(300), outer.winfo_width() or px(420)))
            except Exception:
                pass

        frm.bind("<Configure>", _cap_interests_height)
        _cap_interests_height()

        first_entry.focus_set()
        
        current_main_frame = frm # <-- NEW


    # --- Event Mode: Startup prompt (content row=1) ---
    prompt = ttk.Frame(root, padding=(px(24), px(24)), style="TFrame")
    # Note: This is NOT gridded at startup anymore.
    root.grid_columnconfigure(0, weight=1)
    root.grid_rowconfigure(1, weight=1)

    ttk.Label(prompt, text="Club/Organization:", font=title_font, style="TLabel").grid(
        row=0, column=0, columnspan=2, sticky="w", pady=(0, px(8))
    )
    club_entry = ttk.Entry(prompt, textvariable=club_org_var, width=60, style="CTM.TEntry",
                           validate="key", validatecommand=(root.register(lambda v: len(v) <= 50), "%P"))
    club_entry.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, px(12)))

    ttk.Label(prompt, text="Meeting purpose:", font=title_font, style="TLabel").grid(
        row=2, column=0, columnspan=2, sticky="w", pady=(0, px(8))
    )
    topic_entry = ttk.Entry(prompt, textvariable=meeting_topic_var, width=60, style="CTM.TEntry",
                            validate="key", validatecommand=(root.register(lambda v: len(v) <= 50), "%P"))
    topic_entry.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(0, px(12)))

    ttk.Label(prompt, text="Optional: comma-separated interest buttons (e.g., CAD, 3D Printing, Robotics):",
              style="TLabel").grid(row=4, column=0, columnspan=2, sticky="w", pady=(px(6), px(2)))
    interest_entry = ttk.Entry(prompt, textvariable=interest_input_var, width=60, style="CTM.TEntry")
    interest_entry.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(0, px(10)))

    prompt_status = tk.StringVar(value="")
    ttk.Label(prompt, textvariable=prompt_status, style="TLabel", foreground=BAD_RED).grid(
        row=6, column=0, columnspan=2, sticky="w", pady=(0, px(8))
    )

    def complete_prompt():
        nonlocal interest_labels
        club  = club_org_var.get().strip()
        topic = meeting_topic_var.get().strip()
        if not club:
            prompt_status.set("Please enter the Club/Organization."); return
        if not topic:
            prompt_status.set("Please enter the meeting purpose."); return
        raw = interest_input_var.get().strip()
        interest_labels = [s.strip() for s in raw.split(",") if s.strip()] if raw else []
        root.title(f"{club} Sign-In")
        prompt.grid_forget()
        build_signin_ui()
        set_cursor_hidden(root, HIDE_CURSOR_IN_KIOSK if bool(root.attributes("-fullscreen")) else False)

    ttk.Button(prompt, text="Start", command=complete_prompt, style="CTM.TButton").grid(row=7, column=0, sticky="w")
    prompt.grid_columnconfigure(0, weight=1)
    prompt.grid_columnconfigure(1, weight=1)
    # Bind Return key to the 'Start' button only when this prompt is active
    club_entry.bind("<Return>", lambda e: complete_prompt())
    topic_entry.bind("<Return>", lambda e: complete_prompt())
    interest_entry.bind("<Return>", lambda e: complete_prompt())


    # --- NEW: Setup Mode Function ---
    def run_setup_mode():
        global APP_CLUB_NAME, APP_OFFICER_PIN, OFFICER_MENU_PIN
        nonlocal club_org_var
        
        # 1. Ask for new club name
        new_club = simpledialog.askstring(
            "Setup", "Enter default Club/Organization Name:",
            initialvalue=APP_CLUB_NAME, parent=root
        )
        
        if new_club is None: # User cancelled
            return 
            
        final_club = new_club.strip() if new_club else APP_CLUB_NAME # Keep old if empty

        # 2. Ask for new PIN
        final_pin = APP_OFFICER_PIN # Default to current
        while True:
            new_pin = simpledialog.askstring(
                "Setup", "Enter 4-digit Officer PIN:",
                initialvalue=final_pin, parent=root, show="*" # Show '*' for PIN
            )
            if new_pin is None: # User cancelled
                return
            
            if new_pin.isdigit() and len(new_pin) == 4:
                final_pin = new_pin
                break
            messagebox.showerror("Invalid PIN", "PIN must be exactly 4 digits.", parent=root)

        # 3. Save and update
        try:
            app_config = {"default_club_name": final_club, "officer_pin": final_pin}
            save_config(app_config)
            
            # 4. Update running state
            APP_CLUB_NAME = final_club
            APP_OFFICER_PIN = final_pin
            OFFICER_MENU_PIN = final_pin # Update the global
            club_org_var.set(final_club) # Update the stringvar
            
            messagebox.showinfo("Setup Complete", "Configuration saved.", parent=root)
        
        except Exception as e:
            _excepthook(type(e), e, e.__traceback__)
            messagebox.showerror("Setup Error", f"Failed to save config:\n{e}", parent=root)


    # --- Mode Selection UI ---
    mode_frm = ttk.Frame(root, style="TFrame")
    mode_frm.grid(row=1, column=0, sticky="nsew") # Gridded at startup
    mode_frm.grid_columnconfigure(0, weight=1)
    mode_frm.grid_rowconfigure(1, weight=1)
    mode_frm.grid_rowconfigure(3, weight=1)

    mode_inner = ttk.Frame(mode_frm, style="TFrame", padding=(px(24), px(24)))
    mode_inner.grid(row=2, column=0, sticky="n") # Centered

    ttk.Label(mode_inner, text="Select Sign-In Mode", font=title_font, style="TLabel").pack(pady=(px(20), px(20)))

    event_btn = ttk.Button(
        mode_inner,
        text="Event / Meeting Sign-In",
        style="CTM.TButton",
        width=24,
        command=lambda: (
            mode_frm.grid_forget(),
            prompt.grid(row=1, column=0, sticky="nsew"),
            club_entry.focus_set()
        )
    )
    event_btn.pack(pady=(px(6), px(12)), ipady=px(10))

    makerspace_btn = ttk.Button(
        mode_inner,
        text="MakerSpace / Room Sign-In",
        style="CTM.TButton",
        width=24,
        command=lambda: (
            mode_frm.grid_forget(),
            build_makerspace_ui()
        )
    )
    makerspace_btn.pack(pady=(px(6), px(12)), ipady=px(10))

    # --- NEW: Setup Button ---
    setup_btn = ttk.Button(
        mode_inner,
        text="Setup / Configure",
        style="CTM.TButton",
        width=24,
        command=run_setup_mode # Hook to the new setup function
    )
    setup_btn.pack(pady=(px(6), px(12)), ipady=px(10))

    event_btn.focus_set() # Start with event button focused
    
    # Bind return key to the focused button
    event_btn.bind("<Return>", lambda e: event_btn.invoke())
    makerspace_btn.bind("<Return>", lambda e: makerspace_btn.invoke())
    setup_btn.bind("<Return>", lambda e: setup_btn.invoke())


    # Show the window now that the UI is ready
    root.deiconify()
    root.mainloop()


if __name__ == "__main__":
    main()